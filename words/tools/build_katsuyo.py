#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
活用編（動詞の変化形／形容詞の比較変化）データ生成スクリプト
  入力 : data/source/my_words.xlsx の「7動詞の変化形」「過」「8比較表現」タブ
  出力 : data/katsuyo.js ( window.KATSUYO_WORDS = [...] / window.KATSUYO_META = {...} )

ルール
  - 動詞：不規則・パターン系（「7動詞の変化形」タブ）を全採用 ＋
          規則動詞（「過」タブ・過去形=過去分詞形）を精選し、ちょうど150パターンに調整。
  - 形容詞：「8比較表現」タブを全採用 ＋ 同水準の regular -er/-est・more/most 語を追加し、
           ちょうど50パターンに調整。
  - 1パターン＝3つの疑似単語（原形／過去形(比較級)／過去分詞形(最上級)）として出力。
    words/js/app.js 側で「原形→過去形→過去分詞形」の3連続出題に使う
    （kid でパターンをグループ化し、slot 0→1→2 の順で並べる）。
  - words/data/words.js（基本編/拡張編の2000語）とは別ファイル・別id空間（文字列id）。
    合計 150+50=200パターン × 3点 = 600点満点。

実行 : python3 words/tools/build_katsuyo.py
"""
import openpyxl, re, json, os

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
SRC = os.path.join(ROOT, "data", "source", "my_words.xlsx")
WORDS_JS = os.path.join(ROOT, "data", "words.js")
OUT = os.path.join(ROOT, "data", "katsuyo.js")

TARGET_VERB = 150
TARGET_ADJ = 50

# ---- データ誤りの訂正（教材原本のタイプミス。採点対象なので正しい英語に直す） ----
VERB_FIXES = {
    "plan": {"past": "planned", "pp": "planned"},   # 誤植: planed（plane＝かんなをかけるの過去形）
    "wake": {"past": "woke", "pp": "woken"},          # 誤植: waken（標準は wake-woke-woken）
}
# 定型句寄りで「活用ドリル」の題材としての優先度が低い規則動詞（150に調整するため除外）
DROP_REGULAR_VERBS = {"excuse", "pardon", "kick", "skate", "ski", "spell", "welcome"}
# 「7動詞の変化形」に載っているが原形として不適切（三単現形などの重複）
DROP_IRREGULAR_BASES = {"studies"}

# 「8比較表現」に無い、同水準（JHSレベル・regular -er/-est or more/most）の追加50-38=12語
EXTRA_ADJ = [
    ("kind", "kinder", "kindest", "親切な"),
    ("nice", "nicer", "nicest", "すてきな"),
    ("sad", "sadder", "saddest", "悲しい"),
    ("cute", "cuter", "cutest", "かわいい"),
    ("clean", "cleaner", "cleanest", "きれいな"),
    ("strong", "stronger", "strongest", "強い"),
    ("weak", "weaker", "weakest", "弱い"),
    ("happy", "happier", "happiest", "幸せな"),
    ("sunny", "sunnier", "sunniest", "晴れた"),
    ("careful", "more careful", "most careful", "注意深い"),
    ("dangerous", "more dangerous", "most dangerous", "危険な"),
    ("expensive", "more expensive", "most expensive", "高価な"),
]


def is_word(s):
    if not isinstance(s, str):
        return False
    s = s.strip()
    if not s or s == "-":
        return False
    return bool(re.match(r"^[A-Za-z][A-Za-z()]*$", s))


def clean_form(s):
    """ 'forgot(ten)' -> ('forgotten', 'forgot')。丸括弧が無ければ (s, None)。 """
    s = s.strip()
    m = re.match(r"^([a-zA-Z]+)\(([a-zA-Z]+)\)$", s)
    if m:
        return m.group(1) + m.group(2), m.group(1)
    return s, None


def load_jp_lookup():
    """ words.js の既存語彙から j(意味) を引けるようにする（表記を既存基本語彙と揃える） """
    text = open(WORDS_JS, encoding="utf-8").read()
    arr = json.loads(re.search(r"window\.WORDS\s*=\s*(\[.*\]);", text).group(1))
    m = {}
    for w in arr:
        k = w["w"].strip().lower()
        if k not in m:
            m[k] = w["j"]
    return m


def extract_verbs(wb, jp_lookup):
    ws = wb["7動詞の変化形"]
    irregular, seen = [], set()
    for row in ws.iter_rows(min_row=4, values_only=True):
        for off in (0, 6):    # 左ブロック A-D(0-3) ／ 右ブロック G-J(6-9)。E,F(4,5)は空列
            jp, base, past, pp = row[off], row[off + 1], row[off + 2], row[off + 3]
            if not (is_word(base) and is_word(past) and is_word(pp)):
                continue
            base_l = base.strip().lower()
            if base_l in seen or base_l in DROP_IRREGULAR_BASES or past.strip() == "-":
                continue
            pastf, past_alt = clean_form(past)
            ppf, pp_alt = clean_form(pp)
            jpv = jp.strip() if isinstance(jp, str) and jp.strip() else jp_lookup.get(base_l, "")
            seen.add(base_l)
            fix = VERB_FIXES.get(base_l)
            if fix:
                pastf, ppf = fix["past"], fix["pp"]
                past_alt = pp_alt = None
            irregular.append(dict(base=base.strip(), past=pastf, pp=ppf, jp=jpv,
                                   past_alt=past_alt, pp_alt=pp_alt))

    ws2 = wb["過"]   # 規則動詞の補充プール（過去形＝過去分詞形）。左A-C(0-2)／右F-H(5-7)
    regular = []
    for row in ws2.iter_rows(min_row=4, values_only=True):
        for off in (0, 5):
            jp, base, past = row[off], row[off + 1], row[off + 2]
            if not (is_word(base) and is_word(past)):
                continue
            base_l = base.strip().lower()
            if base_l in seen or past.strip() == "-" or base_l in DROP_REGULAR_VERBS:
                continue
            jpv = jp.strip() if isinstance(jp, str) and jp.strip() else jp_lookup.get(base_l, "")
            seen.add(base_l)
            pastf = past.strip()
            fix = VERB_FIXES.get(base_l)
            if fix:
                pastf = fix["past"]
            regular.append(dict(base=base.strip(), past=pastf, pp=pastf, jp=jpv,
                                 past_alt=None, pp_alt=None))

    need = TARGET_VERB - len(irregular)
    if need < 0 or need > len(regular):
        raise SystemExit(f"想定外：不規則{len(irregular)}語・規則プール{len(regular)}語で{TARGET_VERB}に届かない/超過")
    verbs = irregular + regular[:need]
    if len(verbs) != TARGET_VERB:
        raise SystemExit(f"動詞が{TARGET_VERB}にならない: {len(verbs)}")
    if len(set(v["base"].lower() for v in verbs)) != TARGET_VERB:
        raise SystemExit("動詞baseが重複している")
    return verbs


def extract_adjs(wb):
    ws = wb["8比較表現"]
    adjs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        jp_line, en_line = row[0], row[1]
        if not (isinstance(en_line, str) and en_line.strip()):
            continue
        parts = en_line.strip().split()
        if len(parts) == 3:
            base, comp, sup = parts
        elif len(parts) == 5:              # "useful more useful most useful"
            base = parts[0]
            comp = parts[1] + " " + parts[2]
            sup = parts[3] + " " + parts[4]
        else:
            raise SystemExit(f"8比較表現の形式が想定外: {en_line!r}")
        jp_parts = jp_line.split() if isinstance(jp_line, str) else []
        adjs.append(dict(base=base, comp=comp, sup=sup, jp=(jp_parts[0] if jp_parts else "")))

    existing = {a["base"].lower() for a in adjs}
    for base, comp, sup, jp in EXTRA_ADJ:
        if base.lower() in existing:
            raise SystemExit(f"追加形容詞が既存と重複: {base}")
        adjs.append(dict(base=base, comp=comp, sup=sup, jp=jp))

    if len(adjs) != TARGET_ADJ:
        raise SystemExit(f"形容詞が{TARGET_ADJ}にならない: {len(adjs)}")
    if len(set(a["base"].lower() for a in adjs)) != TARGET_ADJ:
        raise SystemExit("形容詞baseが重複している")
    return adjs


# ---- パターン → 3疑似単語（原形／過去形or比較級／過去分詞形or最上級） ----
VERB_LABELS = ["原形", "過去形", "過去分詞形"]
ADJ_LABELS = ["原形", "比較級", "最上級"]


def make_verb_entries(v, idx):
    kid = f"v{idx}"
    base, past, pp, jp = v["base"], v["past"], v["pp"], v["jp"]
    j0 = jp
    j1 = f"{jp}\n原形：{base}"
    j2 = f"{jp}\n原形：{base}　過去形：{past}"
    out = [
        dict(id=f"{kid}_0", w=base, j=j0, p="動詞の活用", m="k", g=0, kid=kid, slot=0, slotLabel=VERB_LABELS[0]),
        dict(id=f"{kid}_1", w=past, j=j1, p="動詞の活用", m="k", g=0, kid=kid, slot=1, slotLabel=VERB_LABELS[1]),
        dict(id=f"{kid}_2", w=pp, j=j2, p="動詞の活用", m="k", g=0, kid=kid, slot=2, slotLabel=VERB_LABELS[2]),
    ]
    if v.get("past_alt"):
        out[1]["alt"] = [v["past_alt"]]
    if v.get("pp_alt"):
        out[2]["alt"] = [v["pp_alt"]]
    return out


def make_adj_entries(a, idx):
    kid = f"a{idx}"
    base, comp, sup, jp = a["base"], a["comp"], a["sup"], a["jp"]
    j0 = jp
    j1 = f"{jp}\n原形：{base}"
    j2 = f"{jp}\n原形：{base}　比較級：{comp}"
    return [
        dict(id=f"{kid}_0", w=base, j=j0, p="形容詞の比較", m="k", g=0, kid=kid, slot=0, slotLabel=ADJ_LABELS[0]),
        dict(id=f"{kid}_1", w=comp, j=j1, p="形容詞の比較", m="k", g=0, kid=kid, slot=1, slotLabel=ADJ_LABELS[1]),
        dict(id=f"{kid}_2", w=sup, j=j2, p="形容詞の比較", m="k", g=0, kid=kid, slot=2, slotLabel=ADJ_LABELS[2]),
    ]


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    jp_lookup = load_jp_lookup()
    verbs = extract_verbs(wb, jp_lookup)
    adjs = extract_adjs(wb)
    wb.close()

    entries = []
    for i, v in enumerate(verbs, 1):
        entries += make_verb_entries(v, i)
    for i, a in enumerate(adjs, 1):
        entries += make_adj_entries(a, i)

    meta = dict(verbCount=len(verbs), adjCount=len(adjs),
                patternCount=len(verbs) + len(adjs),
                maxScore=(len(verbs) + len(adjs)) * 3)

    with open(OUT, "w", encoding="utf-8") as f:
        f.write("// 自動生成ファイル（words/tools/build_katsuyo.py）。直接編集しないでください。\n")
        f.write("// 動詞の活用（原形/過去形/過去分詞形）＋形容詞の比較変化（原形/比較級/最上級）。\n")
        f.write("window.KATSUYO_META = " + json.dumps(meta, ensure_ascii=False) + ";\n")
        f.write("window.KATSUYO_WORDS = " + json.dumps(entries, ensure_ascii=False, separators=(",", ":")) + ";\n")

    print("動詞パターン:", meta["verbCount"], "×3点=", meta["verbCount"] * 3, "点")
    print("形容詞パターン:", meta["adjCount"], "×3点=", meta["adjCount"] * 3, "点")
    print("合計:", meta["patternCount"], "パターン／満点", meta["maxScore"], "点")
    print("疑似単語エントリ数:", len(entries), "（=パターン数×3）")
    print("->", OUT)


if __name__ == "__main__":
    main()
