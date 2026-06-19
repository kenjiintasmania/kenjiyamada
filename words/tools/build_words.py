#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
単語データ生成スクリプト
  入力 : data/source/my_words.xlsx   … 先生（オレ）の単語集（品詞別タブ）→ 基本編
         data/source/new_horizon.xlsx … NEW HORIZON 単語データベース      → 拡張編
  出力 : data/words.js  ( window.WORDS = [...] / window.WORD_META = {...} )

ルール
  - 基本編 = my_words.xlsx の品詞別タブから抽出（意味は先生の表記を優先）
  - 拡張編 = NEW HORIZON から「基本編にない語」を重要度順に並べ、
             基本編とあわせて合計ちょうど 2000 語になるところまで採用
             （あふれた＝マイナー過ぎる語は除外）
  - 品詞は細分化。代名詞と疑問詞は必ず分離する。
実行 : python3 tools/build_words.py
"""
import openpyxl, re, json, os
from collections import Counter, OrderedDict

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
ANS = os.path.join(ROOT, "data", "source", "my_words.xlsx")
NH  = os.path.join(ROOT, "data", "source", "new_horizon.xlsx")
OUT = os.path.join(ROOT, "data", "words.js")

TARGET = 2000

EN      = re.compile(r"^[A-Za-z][A-Za-z'’\-]*$")        # 単一語（ハイフン・アポストロフィ可）
ENMULTI = re.compile(r"^[A-Za-z][A-Za-z'’\- ]*$")       # 連語（スペース可）
JP      = re.compile(r"[ぁ-んァ-ヶ一-龥ー]")

def norm(w):
    return str(w).strip().replace('’', "'").lower()

# 疑問詞（NHでは代名詞/副詞に混在しているので必ずここへ寄せる）
INTERROG = {"what","who","whom","whose","which","where","when","why","how",
            "what time","how many","how much","how old","how long","how often","how far",
            "what kind","what sport","what color","what colour","what day","what subject"}
PREPS = {"at","on","in","over","under","by","near","around","into","to","for","from","with",
         "of","about","after","before","between","during","without","through","across","along",
         "behind","below","above","up","down","off","out","than"}
CONJ  = {"and","but","so","because","if","when","that","or","while","though","although",
         "as","since","before","after","until"}

# 手動補正（空欄/表記ゆれ）
MANUAL_JP = {"twenty-one":"21", "social":"社会（科）の", "economics":"経済（科）"}
DROP = {"studies"}  # 三人称単数形などのノイズ（原形 study を採用）

POSMAP = {'名':'名詞','動':'動詞','形':'形容詞','副':'副詞','代':'代名詞','前':'前置詞',
          '接':'接続詞','助':'助動詞','冠':'冠詞','間':'間投詞','短':'短縮形','連':'連語',
          '名/形':'名詞','名/副':'名詞','助/動':'助動詞','動/助':'動詞','接/前':'接続詞','前/接':'前置詞'}

def fine_pos(word, raw_pos, default=None):
    if norm(word) in INTERROG:
        return '疑問詞'
    return POSMAP.get(raw_pos, default)

# ---------- NEW HORIZON データベース ----------
def load_nh():
    wb = openpyxl.load_workbook(NH, read_only=True, data_only=True)
    ws = wb["①Word List"]
    m = OrderedDict()
    for r in ws.iter_rows(min_row=4, values_only=True):
        w = r[2]
        if not w:
            continue
        k = norm(w)
        if k in m:
            continue
        m[k] = dict(word=str(w).strip(), pos=r[8], jp=r[9], grade=r[10],
                    sho=(r[3] == '〇'), hassin=(r[4] == '〇'),
                    rengo=(r[5] == '〇'), shoshutsu=(r[6] == '〇'))
    wb.close()
    return m

# ---------- 基本編（先生の品詞別タブ） ----------
def load_basic(NHMAP):
    wb = openpyxl.load_workbook(ANS, read_only=True, data_only=True)
    def cell(ws, row, col):
        v = ws.cell(row=row, column=col).value
        return v.strip() if isinstance(v, str) else v

    basic = OrderedDict()
    def add(word, jp, pos, src, allow_multi=False):
        if word is None:
            return
        word = str(word).strip().replace('’', "'")
        if not word:
            return
        rx = ENMULTI if allow_multi else EN
        if not rx.match(word):
            return
        k = norm(word)
        if k in DROP or k in basic:
            return
        if k in INTERROG:
            pos = '疑問詞'
        if jp is not None and not (isinstance(jp, str) and JP.search(jp)):
            jp = None
        if not jp and k in MANUAL_JP:
            jp = MANUAL_JP[k]
        if not jp and k in NHMAP:
            jp = NHMAP[k]['jp']
        basic[k] = dict(word=word, jp=(jp or ''), pos=pos, src=src)

    def pos_from_label(lbl, default):
        if not isinstance(lbl, str):
            return default
        for ch, p in (('副','副詞'), ('形','形容詞'), ('動','動詞'), ('名','名詞')):
            if ch in lbl:
                return p
        return default

    ws = wb["★1主語と疑問詞"]
    for row in range(2, 60):
        add(cell(ws, row, 2), cell(ws, row, 1), '代名詞', '★1', allow_multi=True)

    ws = wb["★2小学動詞"]
    for row in range(4, 60):
        add(cell(ws, row, 4), cell(ws, row, 2), '動詞', '★2')
        add(cell(ws, row, 9), cell(ws, row, 7), '動詞', '★2')

    ws = wb["★3小学形副"]
    for row in range(3, 80):
        add(cell(ws, row, 4), cell(ws, row, 2), pos_from_label(cell(ws, row, 1), '形容詞'), '★3')
        add(cell(ws, row, 9), cell(ws, row, 7), pos_from_label(cell(ws, row, 6), '形容詞'), '★3')

    ws = wb["4&5名詞"]
    for row in range(3, 200):
        num = cell(ws, row, 2)
        card, ordn = cell(ws, row, 3), cell(ws, row, 4)
        if card:
            add(card, (str(num) + "（つの）" if num is not None else None), '名詞', '4&5')
        if ordn:
            add(ordn, (f"{num}番目" if num is not None else None), '名詞', '4&5')
        for jc, wc in [(6, 7), (9, 10), (12, 13)]:
            w = cell(ws, row, wc)
            if not w:
                continue
            kw = norm(w)
            pos = '疑問詞' if kw in INTERROG else ('前置詞' if kw in PREPS else '名詞')
            add(w, cell(ws, row, jc), pos, '4&5', allow_multi=(kw in INTERROG))

    ws = wb["★6前置詞と接続詞"]
    for row in range(2, 40):
        en = cell(ws, row, 2)
        if en:
            add(en, cell(ws, row, 1), '接続詞' if norm(en) in CONJ else '前置詞', '★6')

    ws = wb["7動詞の変化形"]
    for row in range(4, 120):
        add(cell(ws, row, 2), cell(ws, row, 1), '動詞', '7')
        add(cell(ws, row, 8), cell(ws, row, 7), '動詞', '7')

    ws = wb["進"]
    for row in range(4, 120):
        add(cell(ws, row, 2), cell(ws, row, 1), '動詞', '進')
        add(cell(ws, row, 6), cell(ws, row, 5), '動詞', '進')

    ws = wb["過"]
    for row in range(4, 120):
        add(cell(ws, row, 2), cell(ws, row, 1), '動詞', '過')
        add(cell(ws, row, 7), cell(ws, row, 6), '動詞', '過')

    ws = wb["8比較表現"]
    for row in range(2, 80):
        trip = cell(ws, row, 2)
        jp = cell(ws, row, 1)
        if isinstance(trip, str) and trip.strip():
            base = trip.strip().split()[0]
            add(base, (jp.split()[0] if isinstance(jp, str) else None), '形容詞', '8')
    wb.close()

    # 主語タブの here/there などは NH に従って副詞に補正
    for k, v in basic.items():
        if k in INTERROG:
            v['pos'] = '疑問詞'
        elif v['src'] == '★1' and k in NHMAP and fine_pos(v['word'], NHMAP[k]['pos']) == '副詞':
            v['pos'] = '副詞'
    return basic

# ---------- 拡張編（NH − 基本、重要度順、合計2000で切る） ----------
def build_extended(NHMAP, basic):
    need = TARGET - len(basic)
    bk = set(basic.keys())
    cands = []
    for k, v in NHMAP.items():
        if k in bk:
            continue
        fp = fine_pos(v['word'], v['pos'])
        if fp is None:
            continue
        score = (4 if v['hassin'] else 0) + (3 if v['sho'] else 0) + (1 if v['shoshutsu'] else 0)
        score += {1: 3, 2: 2, 3: 1}.get(v['grade'], 0)
        if v['rengo']:
            score += 1
        if v['pos'] == '連':
            score -= 3      # 連語は単語より優先度を下げる
        if v['pos'] == '短':
            score -= 1
        cands.append((score, v['grade'] or 9, k, fp, v))
    cands.sort(key=lambda x: (-x[0], x[1], x[2]))
    kept = cands[:need]
    dropped = cands[need:]
    return kept, dropped

def main():
    NHMAP = load_nh()
    basic = load_basic(NHMAP)
    kept, dropped = build_extended(NHMAP, basic)

    POS_ORDER = ['名詞','動詞','形容詞','副詞','代名詞','疑問詞','前置詞','接続詞',
                 '助動詞','冠詞','間投詞','短縮形','連語']
    words = []
    wid = 0
    for k, v in basic.items():
        wid += 1
        words.append(dict(id=wid, w=v['word'], j=v['jp'], p=v['pos'], m='b',
                          g=(NHMAP[k]['grade'] if k in NHMAP else 0)))
    for score, g, k, fp, v in kept:
        wid += 1
        words.append(dict(id=wid, w=v['word'], j=v['jp'], p=fp, m='e', g=(v['grade'] or 0)))

    # メタ情報（カテゴリ件数）
    def counts(mode):
        c = Counter(w['p'] for w in words if w['m'] == mode)
        return [[p, c[p]] for p in POS_ORDER if c[p]]
    meta = dict(
        total=len(words),
        target=TARGET,
        posOrder=POS_ORDER,
        basicCount=sum(1 for w in words if w['m'] == 'b'),
        extendedCount=sum(1 for w in words if w['m'] == 'e'),
        basicByPos=counts('b'),
        extendedByPos=counts('e'),
        droppedCount=len(dropped),
    )

    with open(OUT, "w", encoding="utf-8") as f:
        f.write("// 自動生成ファイル（tools/build_words.py）。直接編集しないでください。\n")
        f.write("window.WORD_META = " + json.dumps(meta, ensure_ascii=False) + ";\n")
        f.write("window.WORDS = " + json.dumps(words, ensure_ascii=False, separators=(",", ":")) + ";\n")

    # レポート
    print("基本編 :", meta['basicCount'], "語")
    print("拡張編 :", meta['extendedCount'], "語")
    print("合計   :", meta['total'], "語")
    print("除外(マイナー過ぎ):", meta['droppedCount'], "語")
    print("基本編 品詞別:", dict(meta['basicByPos']))
    print("拡張編 品詞別:", dict(meta['extendedByPos']))
    print("→", OUT)

if __name__ == "__main__":
    main()
